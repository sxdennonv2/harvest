"""Generate the single course workbook for the Excel-to-a-High-Standard lessons.

One workbook, one starter tab per lesson. Each lesson tab holds the worked-example
data set up but with the task left undone, so the learner (or an AI prompt targeting
that tab) can complete it. Adding a lesson is a single entry in LESSONS.
"""
from __future__ import annotations

import json
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

OUT = Path(
    r"C:\Users\dennon\VSCodeProjects\ai_training\teach-buying-analyst\excel\artefacts\excel-course-workbook.xlsx"
)
COURSE_JSON = Path(__file__).with_name("course.json")

HEADER_FILL = PatternFill("solid", fgColor="1A7A4C")
RATE_FILL = PatternFill("solid", fgColor="E7F2EC")
TASK_FILL = PatternFill("solid", fgColor="FFF4CE")
INK = "1A1A1A"
WHITE = "FFFFFF"


@dataclass
class Lesson:
    """One lesson tab: starter data plus a stated task, no solution applied."""

    sheet: str
    title: str
    task: str
    products: list[tuple[str, float]] = field(default_factory=list)
    rate: float = 1.92
    rate_named: bool = False
    landed_header: bool = True
    flat_headers: list[str] = field(default_factory=list)
    flat_rows: list[tuple] = field(default_factory=list)


LESSONS: list[Lesson] = [
    Lesson(
        sheet="References",
        title="Lesson 0001 — Relative & Absolute references",
        task=(
            "In C2 type  =B2*$F$1  then fill down to C11. "
            "Change the rate in F1 and watch every row recalculate."
        ),
        landed_header=True,
        products=[
            ("Stacked Chips 12pk 300g", 14.5),
            ("Noodle Snaks 12pk 250g", 11.2),
            ("Rice Crackers 24pk 100g", 18.75),
            ("Salted Pretzels 16pk 200g", 9.8),
            ("Corn Puffs 12pk 150g", 12.4),
            ("Veggie Straws 12pk 120g", 13.6),
            ("Popcorn 10pk 100g", 8.5),
            ("Cheese Twists 16pk 180g", 10.9),
            ("Pretzel Bites 20pk 220g", 15.3),
            ("Tortilla Rounds 12pk 330g", 16.2),
        ],
    ),
    Lesson(
        sheet="Tables",
        title="Lesson 0002 — Tables: the modern foundation",
        task=(
            "Click the data, press Ctrl+T, name it Products on the Table Design tab. "
            "Add a Landed cost column:  =[@[Case cost (GBP)]]*Rate  — it fills down by itself."
        ),
        rate_named=True,
        landed_header=False,
        products=[
            ("Cheddar Cheese 500g", 24.5),
            ("Sourdough Loaf 800g", 18.75),
            ("Free-range Eggs 12pk", 12.2),
            ("Salted Butter 250g", 15.9),
            ("Greek Yoghurt 1kg", 6.4),
            ("Orange Juice 2L", 4.8),
            ("Chicken Breast 1kg", 11.5),
            ("Basmati Rice 5kg", 13.2),
            ("Olive Oil 1L", 9.75),
            ("Dark Chocolate 200g", 3.9),
        ],
    ),
    Lesson(
        sheet="PivotTable",
        title="Lesson 0003 — PivotTables: summarise in seconds",
        task=(
            "Click the data, then Insert > PivotTable. Drag Category to Rows and "
            "Cases + Revenue (AUD) to Values to total by category. Then add Supplier "
            "under Category in Rows for a breakdown."
        ),
        flat_headers=["Month", "Category", "Supplier", "Product", "Cases", "Revenue (AUD)"],
        flat_rows=[
            ("Jan", "Snacks", "Acme Foods", "Stacked Chips 12pk 300g", 120, 3480.0),
            ("Jan", "Snacks", "Acme Foods", "Corn Puffs 12pk 150g", 90, 2232.0),
            ("Jan", "Bakery", "BrightHarvest", "Sourdough Loaf 800g", 60, 2160.0),
            ("Jan", "Dairy", "Coastal Co", "Cheddar Cheese 500g", 45, 2205.0),
            ("Jan", "Produce", "GreenField", "Free-range Eggs 12pk", 80, 1952.0),
            ("Feb", "Snacks", "Acme Foods", "Stacked Chips 12pk 300g", 140, 4060.0),
            ("Feb", "Snacks", "Harvest Lane", "Rice Crackers 24pk 100g", 70, 2625.0),
            ("Feb", "Bakery", "BrightHarvest", "Sourdough Loaf 800g", 75, 2700.0),
            ("Feb", "Dairy", "Coastal Co", "Salted Butter 250g", 55, 1749.0),
            ("Feb", "Dairy", "Coastal Co", "Greek Yoghurt 1kg", 65, 1248.0),
            ("Feb", "Produce", "GreenField", "Free-range Eggs 12pk", 95, 2318.0),
            ("Mar", "Snacks", "Acme Foods", "Corn Puffs 12pk 150g", 110, 2728.0),
            ("Mar", "Snacks", "Harvest Lane", "Popcorn 10pk 100g", 130, 2210.0),
            ("Mar", "Bakery", "BrightHarvest", "Sourdough Loaf 800g", 85, 3060.0),
            ("Mar", "Dairy", "Coastal Co", "Cheddar Cheese 500g", 50, 2450.0),
            ("Mar", "Produce", "GreenField", "Basmati Rice 5kg", 40, 1056.0),
            ("Mar", "Produce", "GreenField", "Orange Juice 2L", 100, 1920.0),
            ("Mar", "Snacks", "Harvest Lane", "Pretzel Bites 20pk 220g", 60, 1836.0),
        ],
    ),
]


def _header_cell(cell) -> None:
    cell.font = Font(bold=True, color=WHITE)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="left")


def write_flat(ws: Worksheet, lesson: Lesson) -> None:
    """Write a flat, pivot-ready transaction list with a task note beside it."""
    for c, head in enumerate(lesson.flat_headers, start=1):
        cell = ws.cell(1, c, head)
        _header_cell(cell)
    for r, row in enumerate(lesson.flat_rows, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(r, c, val)
            if isinstance(val, float):
                cell.number_format = "#,##0.00"
    widths = {"A": 10, "B": 12, "C": 15, "D": 26, "E": 9, "F": 15}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws["H1"] = "Your task"
    ws["H1"].font = Font(bold=True, color="9A6A00")
    ws["H2"] = lesson.task
    ws["H2"].fill = TASK_FILL
    ws["H2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("H2:L7")
    ws.column_dimensions["H"].width = 16
    for col in ("I", "J", "K", "L"):
        ws.column_dimensions[col].width = 12
    ws.freeze_panes = "A2"


def write_lesson(ws: Worksheet, lesson: Lesson) -> None:
    if lesson.flat_headers:
        write_flat(ws, lesson)
        return
    ws["A1"] = "Product"
    ws["B1"] = "Case cost (GBP)"
    _header_cell(ws["A1"])
    _header_cell(ws["B1"])
    if lesson.landed_header:
        ws["C1"] = "Landed cost (AUD)"
        _header_cell(ws["C1"])

    rate_label = "Rate" if lesson.rate_named else "Exchange rate (GBP to AUD)"
    ws["E1"] = rate_label
    ws["F1"] = lesson.rate
    ws["E1"].font = Font(bold=True)
    ws["F1"].fill = RATE_FILL
    ws["F1"].font = Font(bold=True)
    ws["F1"].number_format = "0.00"

    for i, (name, cost) in enumerate(lesson.products, start=2):
        ws.cell(i, 1, name)
        c = ws.cell(i, 2, cost)
        c.number_format = "0.00"

    ws["E3"] = "Your task"
    ws["E3"].font = Font(bold=True, color="9A6A00")
    ws["E4"] = lesson.task
    ws["E4"].fill = TASK_FILL
    ws["E4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("E4:H8")

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["E"].width = 14
    for col in ("F", "G", "H"):
        ws.column_dimensions[col].width = 12
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Additional lessons (0004–0009): each builds a starter tab and a hidden
# worked-solution tab. Kept as focused functions rather than the Lesson
# dataclass because their layouts differ (summaries, lookups, charts, print).
# ---------------------------------------------------------------------------

SALES_HEADERS = ["Month", "Category", "Supplier", "Product", "Cases", "Revenue (AUD)"]
SALES_ROWS: list[tuple] = [
    ("Jan", "Snacks", "Acme Foods", "Stacked Chips 12pk 300g", 120, 3480.0),
    ("Jan", "Snacks", "Acme Foods", "Corn Puffs 12pk 150g", 90, 2232.0),
    ("Jan", "Bakery", "BrightHarvest", "Sourdough Loaf 800g", 60, 2160.0),
    ("Jan", "Dairy", "Coastal Co", "Cheddar Cheese 500g", 45, 2205.0),
    ("Jan", "Produce", "GreenField", "Free-range Eggs 12pk", 80, 1952.0),
    ("Feb", "Snacks", "Acme Foods", "Stacked Chips 12pk 300g", 140, 4060.0),
    ("Feb", "Snacks", "Harvest Lane", "Rice Crackers 24pk 100g", 70, 2625.0),
    ("Feb", "Bakery", "BrightHarvest", "Sourdough Loaf 800g", 75, 2700.0),
    ("Feb", "Dairy", "Coastal Co", "Salted Butter 250g", 55, 1749.0),
    ("Feb", "Dairy", "Coastal Co", "Greek Yoghurt 1kg", 65, 1248.0),
    ("Feb", "Produce", "GreenField", "Free-range Eggs 12pk", 95, 2318.0),
    ("Mar", "Snacks", "Acme Foods", "Corn Puffs 12pk 150g", 110, 2728.0),
    ("Mar", "Snacks", "Harvest Lane", "Popcorn 10pk 100g", 130, 2210.0),
    ("Mar", "Bakery", "BrightHarvest", "Sourdough Loaf 800g", 85, 3060.0),
    ("Mar", "Dairy", "Coastal Co", "Cheddar Cheese 500g", 50, 2450.0),
    ("Mar", "Produce", "GreenField", "Basmati Rice 5kg", 40, 1056.0),
    ("Mar", "Produce", "GreenField", "Orange Juice 2L", 100, 1920.0),
    ("Mar", "Snacks", "Harvest Lane", "Pretzel Bites 20pk 220g", 60, 1836.0),
]
CATEGORIES = ["Snacks", "Bakery", "Dairy", "Produce"]


def _course_catalogue() -> list[tuple[str, str, str]]:
    """Ordered (number, title, sheet) from course.json — the single source of truth.

    Numbers are derived from list position (1-based, zero-padded); a subcard takes
    its parent's number with a ``.k`` suffix and is placed immediately after it.
    """
    data = json.loads(COURSE_JSON.read_text(encoding="utf-8"))
    catalogue: list[tuple[str, str, str]] = []
    for index, lesson in enumerate(data["lessons"], start=1):
        number = f"{index:04d}"
        catalogue.append((number, lesson["title"], lesson["workbook_sheet"]))
        for k, sub in enumerate(lesson.get("subcards", []), start=1):
            catalogue.append((f"{number}.{k}", sub["title"], sub["workbook_sheet"]))
    return catalogue



def _task_note(ws: Worksheet, label_cell: str, note_cell: str, merge: str, text: str) -> None:
    """Write the standard yellow 'Your task' note used on every starter tab."""
    ws[label_cell] = "Your task"
    ws[label_cell].font = Font(bold=True, color="9A6A00")
    ws[note_cell] = text
    ws[note_cell].fill = TASK_FILL
    ws[note_cell].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(merge)


def _write_sales_list(ws: Worksheet) -> int:
    """Write the shared flat sales list and return the last data row."""
    for c, head in enumerate(SALES_HEADERS, start=1):
        _header_cell(ws.cell(1, c, head))
    for r, row in enumerate(SALES_ROWS, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(r, c, val)
            if isinstance(val, float):
                cell.number_format = "#,##0.00"
    widths = {"A": 10, "B": 12, "C": 15, "D": 26, "E": 9, "F": 15}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    return 1 + len(SALES_ROWS)


def _add_solution_sheet(wb: openpyxl.Workbook, name: str) -> Worksheet:
    """Create a hidden solution tab next to its starter tab."""
    ws = wb.create_sheet(name)
    ws.sheet_state = "hidden"
    return ws


def build_l03_1_slicers(wb: openpyxl.Workbook) -> None:
    """Slicers: reuse the L03 sales list, then add visible slicer/timeline filters.

    openpyxl cannot author live slicers or timelines, so this tab supplies the same
    pivot-ready data plus a task note that walks the learner through inserting them.
    """
    task = (
        "Build the pivot from the PivotTable tab first (Category in Rows, Cases + Revenue in Values). "
        "Then PivotTable Analyze > Insert Slicer, tick Category and Supplier. Add "
        "PivotTable Analyze > Insert Timeline > Month. Drag the panels beside the pivot; "
        "click a button to filter and drag the timeline to one month — that arrangement is the report."
    )
    ws = wb.create_sheet("Slicers")
    _write_sales_list(ws)
    _task_note(ws, "H1", "H2", "H2:L7", task)
    ws.column_dimensions["H"].width = 16
    for col in ("I", "J", "K", "L"):
        ws.column_dimensions[col].width = 12


def build_l04_sumifs(wb: openpyxl.Workbook) -> None:
    """SUMIFS: total Cases and Revenue per category with a criteria-driven sum."""
    task = (
        "In I2 total Cases for Snacks with  =SUMIFS($E$2:$E$19,$B$2:$B$19,H2)  and in J2 "
        "total Revenue with  =SUMIFS($F$2:$F$19,$B$2:$B$19,H2) . Fill down to row 5 — each "
        "category totals itself. The dollar signs lock the list so the fill stays honest."
    )
    for solution in (False, True):
        ws = (
            _add_solution_sheet(wb, "SUMIFS soln")
            if solution
            else wb.create_sheet("SUMIFS")
        )
        last = _write_sales_list(ws)
        _header_cell(ws.cell(1, 8, "Category"))
        _header_cell(ws.cell(1, 9, "Total Cases"))
        _header_cell(ws.cell(1, 10, "Total Revenue (AUD)"))
        for i, cat in enumerate(CATEGORIES, start=2):
            ws.cell(i, 8, cat)
            if solution:
                ws.cell(i, 9, f"=SUMIFS($E$2:$E${last},$B$2:$B${last},H{i})")
                rev = ws.cell(i, 10, f"=SUMIFS($F$2:$F${last},$B$2:$B${last},H{i})")
                rev.number_format = "#,##0.00"
        ws.column_dimensions["H"].width = 12
        ws.column_dimensions["I"].width = 12
        ws.column_dimensions["J"].width = 20
        if not solution:
            _task_note(ws, "H7", "H8", "H8:J12", task)


def build_l05_xlookup(wb: openpyxl.Workbook) -> None:
    """XLOOKUP: pull a supplier and case cost from a catalogue by product."""
    catalogue = [
        ("Stacked Chips 12pk 300g", "Acme Foods", 14.5),
        ("Corn Puffs 12pk 150g", "Acme Foods", 12.4),
        ("Rice Crackers 24pk 100g", "Harvest Lane", 18.75),
        ("Popcorn 10pk 100g", "Harvest Lane", 8.5),
        ("Pretzel Bites 20pk 220g", "Harvest Lane", 15.3),
        ("Sourdough Loaf 800g", "BrightHarvest", 18.75),
        ("Cheddar Cheese 500g", "Coastal Co", 24.5),
        ("Salted Butter 250g", "Coastal Co", 15.9),
        ("Greek Yoghurt 1kg", "Coastal Co", 6.4),
        ("Free-range Eggs 12pk", "GreenField", 12.2),
    ]
    orders = [
        "Cheddar Cheese 500g",
        "Popcorn 10pk 100g",
        "Sourdough Loaf 800g",
        "Free-range Eggs 12pk",
        "Rice Crackers 24pk 100g",
    ]
    task = (
        "In G2 find the supplier with  =XLOOKUP(F2,$A$2:$A$11,$B$2:$B$11)  and in H2 the case "
        "cost with  =XLOOKUP(F2,$A$2:$A$11,$C$2:$C$11) . Fill down. XLOOKUP returns #N/A cleanly "
        "when a product is missing — add a 4th argument like \"not found\" to handle that."
    )
    for solution in (False, True):
        ws = (
            _add_solution_sheet(wb, "XLOOKUP soln")
            if solution
            else wb.create_sheet("XLOOKUP")
        )
        _header_cell(ws.cell(1, 1, "Product"))
        _header_cell(ws.cell(1, 2, "Supplier"))
        _header_cell(ws.cell(1, 3, "Case cost (GBP)"))
        for r, (name, supplier, cost) in enumerate(catalogue, start=2):
            ws.cell(r, 1, name)
            ws.cell(r, 2, supplier)
            ws.cell(r, 3, cost).number_format = "0.00"
        _header_cell(ws.cell(1, 6, "Order product"))
        _header_cell(ws.cell(1, 7, "Supplier"))
        _header_cell(ws.cell(1, 8, "Case cost (GBP)"))
        for r, name in enumerate(orders, start=2):
            ws.cell(r, 6, name)
            if solution:
                ws.cell(r, 7, f"=XLOOKUP(F{r},$A$2:$A$11,$B$2:$B$11)")
                ws.cell(r, 8, f"=XLOOKUP(F{r},$A$2:$A$11,$C$2:$C$11)").number_format = "0.00"
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["F"].width = 26
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 15
        ws.freeze_panes = "A2"
        if not solution:
            _task_note(ws, "F8", "F9", "F9:H13", task)


def build_l06_conditional(wb: openpyxl.Workbook) -> None:
    """Conditional formatting: a colour scale plus a below-target highlight."""
    rows = [
        ("Stacked Chips 12pk 300g", 0.31),
        ("Corn Puffs 12pk 150g", 0.18),
        ("Rice Crackers 24pk 100g", 0.42),
        ("Popcorn 10pk 100g", 0.12),
        ("Sourdough Loaf 800g", 0.27),
        ("Cheddar Cheese 500g", 0.38),
        ("Salted Butter 250g", 0.22),
        ("Greek Yoghurt 1kg", 0.09),
        ("Free-range Eggs 12pk", 0.16),
        ("Basmati Rice 5kg", 0.34),
    ]
    task = (
        "Select B2:B11. Home > Conditional Formatting > Colour Scales for a green-to-red heat "
        "map, then add a rule: Highlight Cells Rules > Less Than > 0.20 with a red fill. "
        "Low-margin lines now jump out without you reading a single number."
    )
    for solution in (False, True):
        sheet = "Conditional format soln" if solution else "Conditional format"
        ws = _add_solution_sheet(wb, sheet) if solution else wb.create_sheet(sheet)
        _header_cell(ws.cell(1, 1, "Product"))
        _header_cell(ws.cell(1, 2, "Margin %"))
        for r, (name, margin) in enumerate(rows, start=2):
            ws.cell(r, 1, name)
            ws.cell(r, 2, margin).number_format = "0%"
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 12
        ws.freeze_panes = "A2"
        if solution:
            ws.conditional_formatting.add(
                "B2:B11",
                ColorScaleRule(
                    start_type="min",
                    start_color="F8696B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="63BE7B",
                ),
            )
            ws.conditional_formatting.add(
                "B2:B11",
                CellIsRule(
                    operator="lessThan",
                    formula=["0.2"],
                    fill=PatternFill("solid", fgColor="FFC7CE"),
                ),
            )
        else:
            _task_note(ws, "D2", "D3", "D3:G7", task)
            ws.column_dimensions["D"].width = 12
            for col in ("E", "F", "G"):
                ws.column_dimensions[col].width = 12


def build_l07_chart(wb: openpyxl.Workbook) -> None:
    """Charts: a clean column chart of revenue by category."""
    data = [("Snacks", 21099.0), ("Bakery", 7920.0), ("Dairy", 7652.0), ("Produce", 8246.0)]
    task = (
        "Select A1:B5, then Insert > Column Chart (2-D clustered). Give it a clear title "
        "(\"Revenue by category\"), delete the legend (one series needs none) and keep the "
        "gridlines light. A clean chart says one thing well."
    )
    for solution in (False, True):
        sheet = "Chart soln" if solution else "Chart"
        ws = _add_solution_sheet(wb, sheet) if solution else wb.create_sheet(sheet)
        _header_cell(ws.cell(1, 1, "Category"))
        _header_cell(ws.cell(1, 2, "Revenue (AUD)"))
        for r, (cat, rev) in enumerate(data, start=2):
            ws.cell(r, 1, cat)
            ws.cell(r, 2, rev).number_format = "#,##0"
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 15
        ws.freeze_panes = "A2"
        if solution:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Revenue by category"
            chart.legend = None
            chart.y_axis.title = "Revenue (AUD)"
            chart.height = 7.5
            chart.width = 13
            values = Reference(ws, min_col=2, min_row=1, max_row=1 + len(data))
            cats = Reference(ws, min_col=1, min_row=2, max_row=1 + len(data))
            chart.add_data(values, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, "D2")
        else:
            _task_note(ws, "D2", "D3", "D3:H8", task)
            ws.column_dimensions["D"].width = 12
            for col in ("E", "F", "G", "H"):
                ws.column_dimensions[col].width = 11


def build_l08_flashfill(wb: openpyxl.Workbook) -> None:
    """Flash Fill & Text to Columns: split names and parse product codes."""
    contacts = [
        ("Bianchi, Marco", "Marco", "Bianchi"),
        ("Okafor, Ada", "Ada", "Okafor"),
        ("Nguyen, Linh", "Linh", "Nguyen"),
        ("Smith, Jordan", "Jordan", "Smith"),
        ("Patel, Riya", "Riya", "Patel"),
    ]
    codes = [
        ("SNAK-001-300G", "SNAK", "001", "300G"),
        ("BAKE-014-800G", "BAKE", "014", "800G"),
        ("DARY-007-500G", "DARY", "007", "500G"),
        ("PROD-022-1KG", "PROD", "022", "1KG"),
        ("SNAK-009-150G", "SNAK", "009", "150G"),
    ]
    task = (
        "Flash Fill: C2 already shows the first name — start typing C3 and press Ctrl+E to fill "
        "the rest; repeat for Last name in D. Text to Columns: select F2:F6, Data > Text to "
        "Columns > Delimited > Other '-' to split code, number and size into G, H, I."
    )
    for solution in (False, True):
        sheet = "Flash Fill soln" if solution else "Flash Fill"
        ws = _add_solution_sheet(wb, sheet) if solution else wb.create_sheet(sheet)
        _header_cell(ws.cell(1, 1, "Contact (Surname, First)"))
        _header_cell(ws.cell(1, 3, "First name"))
        _header_cell(ws.cell(1, 4, "Last name"))
        for r, (raw, first, last) in enumerate(contacts, start=2):
            ws.cell(r, 1, raw)
            if solution or r == 2:
                ws.cell(r, 3, first)
                ws.cell(r, 4, last)
        _header_cell(ws.cell(1, 6, "Product code"))
        _header_cell(ws.cell(1, 7, "Dept"))
        _header_cell(ws.cell(1, 8, "Number"))
        _header_cell(ws.cell(1, 9, "Size"))
        for r, (code, dept, num, size) in enumerate(codes, start=2):
            ws.cell(r, 6, code)
            if solution:
                ws.cell(r, 7, dept)
                ws.cell(r, 8, num)
                ws.cell(r, 9, size)
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["C"].width = 13
        ws.column_dimensions["D"].width = 13
        ws.column_dimensions["F"].width = 16
        for col in ("G", "H", "I"):
            ws.column_dimensions[col].width = 10
        ws.freeze_panes = "A2"
        if not solution:
            _task_note(ws, "A8", "A9", "A9:D13", task)


def build_l09_print(wb: openpyxl.Workbook) -> None:
    """Print- & export-ready: a tidy table set up to print on one clean page."""
    rows = [
        ("Stacked Chips 12pk 300g", "Snacks", "Acme Foods", 14.5, 260, 7540.0),
        ("Corn Puffs 12pk 150g", "Snacks", "Acme Foods", 12.4, 200, 4960.0),
        ("Rice Crackers 24pk 100g", "Snacks", "Harvest Lane", 18.75, 140, 5250.0),
        ("Popcorn 10pk 100g", "Snacks", "Harvest Lane", 8.5, 260, 4420.0),
        ("Sourdough Loaf 800g", "Bakery", "BrightHarvest", 18.75, 220, 7920.0),
        ("Cheddar Cheese 500g", "Dairy", "Coastal Co", 24.5, 95, 4655.0),
        ("Salted Butter 250g", "Dairy", "Coastal Co", 15.9, 55, 1749.0),
        ("Greek Yoghurt 1kg", "Dairy", "Coastal Co", 6.4, 65, 1248.0),
        ("Free-range Eggs 12pk", "Produce", "GreenField", 12.2, 175, 4270.0),
        ("Basmati Rice 5kg", "Produce", "GreenField", 13.2, 40, 1056.0),
        ("Orange Juice 2L", "Produce", "GreenField", 4.8, 100, 1920.0),
        ("Dark Chocolate 200g", "Snacks", "Acme Foods", 3.9, 130, 2340.0),
    ]
    headers = ["Product", "Category", "Supplier", "Case cost (GBP)", "Cases", "Revenue (AUD)"]
    task = (
        "Set the range as the print area (Page Layout > Print Area), choose Landscape and "
        "Fit Sheet on One Page, repeat row 1 on every page (Print Titles > Rows to repeat), and "
        "add a footer with the page number. File > Print to preview a clean, one-page report."
    )
    for solution in (False, True):
        sheet = "Print soln" if solution else "Print ready"
        ws = _add_solution_sheet(wb, sheet) if solution else wb.create_sheet(sheet)
        for c, head in enumerate(headers, start=1):
            _header_cell(ws.cell(1, c, head))
        for r, row in enumerate(rows, start=2):
            for c, val in enumerate(row, start=1):
                cell = ws.cell(r, c, val)
                if isinstance(val, float):
                    cell.number_format = "#,##0.00"
        widths = {"A": 26, "B": 12, "C": 15, "D": 15, "E": 9, "F": 15}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
        last = 1 + len(rows)
        if solution:
            ws.print_area = f"A1:F{last}"
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
            ws.print_title_rows = "1:1"
            ws.oddFooter.center.text = "Buying report — page &P of &N"
            ws.print_options.horizontalCentered = True
        else:
            _task_note(ws, "H1", "H2", "H2:K6", task)
            ws.column_dimensions["H"].width = 14
            for col in ("I", "J", "K"):
                ws.column_dimensions[col].width = 12


def build_sort_filter(wb: openpyxl.Workbook) -> None:
    """Sort & Filter: the shared sales list, ready to order and isolate.

    The starter keeps the raw entry order; the solution shows it sorted by Revenue
    descending with the filter arrows switched on (openpyxl can set the auto-filter
    range but not a live applied filter, so the arrangement demonstrates the outcome).
    """
    task = (
        "Press Ctrl+Shift+L to switch on the filter arrows. Open the Revenue (AUD) dropdown "
        "and Sort Largest to Smallest, then open the Category dropdown and tick only Snacks. "
        "Clear the filter to bring every row back — filtering hides, it never deletes."
    )
    for solution in (False, True):
        ws = (
            _add_solution_sheet(wb, "Sort & Filter soln")
            if solution
            else wb.create_sheet("Sort & Filter")
        )
        rows = sorted(SALES_ROWS, key=lambda r: r[5], reverse=True) if solution else SALES_ROWS
        for c, head in enumerate(SALES_HEADERS, start=1):
            _header_cell(ws.cell(1, c, head))
        for r, row in enumerate(rows, start=2):
            for c, val in enumerate(row, start=1):
                cell = ws.cell(r, c, val)
                if isinstance(val, float):
                    cell.number_format = "#,##0.00"
        widths = {"A": 10, "B": 12, "C": 15, "D": 26, "E": 9, "F": 15}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
        last = 1 + len(rows)
        if solution:
            ws.auto_filter.ref = f"A1:F{last}"
        else:
            _task_note(ws, "H1", "H2", "H2:L6", task)
            ws.column_dimensions["H"].width = 16
            for col in ("I", "J", "K", "L"):
                ws.column_dimensions[col].width = 12


def build_data_validation(wb: openpyxl.Workbook) -> None:
    """Data validation: an order-entry block plus approved lists for dropdowns."""
    approved_cats = ["Snacks", "Bakery", "Dairy", "Produce"]
    approved_suppliers = [
        "Acme Foods",
        "Harvest Lane",
        "BrightHarvest",
        "Coastal Co",
        "GreenField",
        "Riverside",
    ]
    entry_headers = ["Date", "Category", "Supplier", "Product", "Cases"]
    sample = [
        ("2026-01-05", "Snacks", "Acme Foods", "Stacked Chips 12pk 300g", 120),
        ("2026-01-06", "Bakery", "BrightHarvest", "Sourdough Loaf 800g", 60),
        ("2026-01-07", "Dairy", "Coastal Co", "Cheddar Cheese 500g", 45),
    ]
    task = (
        "Select B2:B11, then Data > Data Validation > Allow: List, Source: =$H$2:$H$5 for the "
        "approved categories. Do the same for C2:C11 against =$I$2:$I$7. Each cell now offers a "
        "dropdown and rejects anything off the list — one setup protects every row you type."
    )
    for solution in (False, True):
        ws = (
            _add_solution_sheet(wb, "Data validation soln")
            if solution
            else wb.create_sheet("Data validation")
        )
        for c, head in enumerate(entry_headers, start=1):
            _header_cell(ws.cell(1, c, head))
        for r, row in enumerate(sample, start=2):
            for c, val in enumerate(row, start=1):
                ws.cell(r, c, val)
        _header_cell(ws.cell(1, 8, "Approved categories"))
        for r, cat in enumerate(approved_cats, start=2):
            ws.cell(r, 8, cat)
        _header_cell(ws.cell(1, 9, "Approved suppliers"))
        for r, sup in enumerate(approved_suppliers, start=2):
            ws.cell(r, 9, sup)
        widths = {"A": 12, "B": 12, "C": 15, "D": 26, "E": 9, "H": 18, "I": 15}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
        if solution:
            dv_cat = DataValidation(type="list", formula1="=$H$2:$H$5", allow_blank=True)
            dv_sup = DataValidation(type="list", formula1="=$I$2:$I$7", allow_blank=True)
            ws.add_data_validation(dv_cat)
            dv_cat.add("B2:B11")
            ws.add_data_validation(dv_sup)
            dv_sup.add("C2:C11")
        else:
            _task_note(ws, "A13", "A14", "A14:E17", task)


def build_logical_iferror(wb: openpyxl.Workbook) -> None:
    """Logical & IFERROR: a verdict column and an error-proof per-case calc.

    One row has blank Cases on purpose so IFERROR has something to catch.
    """
    logic_rows = [
        ("Stacked Chips 12pk 300g", 120, 3480.0, 120, 100),
        ("Corn Puffs 12pk 150g", 90, 2232.0, 90, 110),
        ("Rice Crackers 24pk 100g", 70, 2625.0, 70, 60),
        ("Popcorn 10pk 100g", 130, 2210.0, 130, 150),
        ("Sourdough Loaf 800g", 85, 3060.0, 85, 80),
        ("Cheddar Cheese 500g", None, 2450.0, 50, 60),
        ("Salted Butter 250g", 55, 1749.0, 55, 50),
        ("Greek Yoghurt 1kg", 65, 1248.0, 65, 70),
        ("Free-range Eggs 12pk", 95, 2318.0, 95, 90),
        ("Basmati Rice 5kg", 40, 1056.0, 40, 45),
    ]
    headers = ["Product", "Cases", "Revenue (AUD)", "Actual", "Target", "Verdict", "Revenue per case"]
    task = (
        "In F2 type  =IF(D2>=E2,\"On target\",\"Below\")  and fill down. In G2 type "
        "=IFERROR(C2/B2,0)  and fill down. Row 7 has blank Cases on purpose — watch IFERROR "
        "return 0 there instead of #DIV/0!."
    )
    for solution in (False, True):
        ws = (
            _add_solution_sheet(wb, "Logical & IFERROR soln")
            if solution
            else wb.create_sheet("Logical & IFERROR")
        )
        for c, head in enumerate(headers, start=1):
            _header_cell(ws.cell(1, c, head))
        for r, (name, cases, revenue, actual, target) in enumerate(logic_rows, start=2):
            ws.cell(r, 1, name)
            if cases is not None:
                ws.cell(r, 2, cases)
            ws.cell(r, 3, revenue).number_format = "#,##0.00"
            ws.cell(r, 4, actual)
            ws.cell(r, 5, target)
            if solution:
                ws.cell(r, 6, f'=IF(D{r}>=E{r},"On target","Below")')
                ws.cell(r, 7, f"=IFERROR(C{r}/B{r},0)").number_format = "#,##0.00"
        widths = {"A": 26, "B": 8, "C": 15, "D": 9, "E": 9, "F": 12, "G": 16}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
        if not solution:
            _task_note(ws, "A13", "A14", "A14:E17", task)


def build_number_formatting(wb: openpyxl.Workbook) -> None:
    """Number formatting: raw values in the starter, formatted (not changed) in the solution."""
    fmt_rows = [
        ("Stacked Chips 12pk 300g", 3480.0, 0.31, date(2026, 1, 15), 480),
        ("Corn Puffs 12pk 150g", 2232.0, 0.18, date(2026, 1, 15), -120),
        ("Rice Crackers 24pk 100g", 2625.0, 0.42, date(2026, 2, 15), 310),
        ("Popcorn 10pk 100g", 2210.0, 0.12, date(2026, 2, 15), -260),
        ("Sourdough Loaf 800g", 3060.0, 0.27, date(2026, 3, 15), 200),
        ("Cheddar Cheese 500g", 2450.0, 0.38, date(2026, 3, 15), 150),
        ("Salted Butter 250g", 1749.0, 0.22, date(2026, 3, 15), -80),
        ("Greek Yoghurt 1kg", 1248.0, 0.09, date(2026, 4, 15), -140),
        ("Free-range Eggs 12pk", 2318.0, 0.16, date(2026, 4, 15), 90),
        ("Basmati Rice 5kg", 1056.0, 0.34, date(2026, 4, 15), 45),
    ]
    headers = ["Product", "Revenue (AUD)", "Margin", "Date", "Variance"]
    task = (
        "Press Ctrl+1 on each column and format, without changing any value: Revenue as currency "
        "($#,##0), Margin as a percentage (0.0%), Date as a short date, and Variance with the "
        "custom code  #,##0;[Red](#,##0)  so losses show red and bracketed. The formula bar still "
        "shows the true number underneath."
    )
    for solution in (False, True):
        ws = (
            _add_solution_sheet(wb, "Number formatting soln")
            if solution
            else wb.create_sheet("Number formatting")
        )
        for c, head in enumerate(headers, start=1):
            _header_cell(ws.cell(1, c, head))
        for r, (name, revenue, margin, day, variance) in enumerate(fmt_rows, start=2):
            ws.cell(r, 1, name)
            rev = ws.cell(r, 2, revenue)
            marg = ws.cell(r, 3, margin)
            dat = ws.cell(r, 4, day)
            var = ws.cell(r, 5, variance)
            if solution:
                rev.number_format = "$#,##0"
                marg.number_format = "0.0%"
                dat.number_format = "dd mmm yyyy"
                var.number_format = "#,##0;[Red](#,##0)"
        widths = {"A": 26, "B": 15, "C": 10, "D": 14, "E": 12}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
        if not solution:
            _task_note(ws, "G1", "G2", "G2:K6", task)
            for col in ("G", "H", "I", "J", "K"):
                ws.column_dimensions[col].width = 12


def build_text_formulas(wb: openpyxl.Workbook) -> None:
    """Text formulas: a messy extract to trim, case and combine into a label."""
    raw_rows = [
        ("  stacked chips 12pk 300g ", "AF-001", "jane.doe@acme.example"),
        (" CORN PUFFS 12pk 150g", "HL-204", "sam.lee@harvest.example"),
        ("rice crackers  24pk 100g ", "BH-330", "amy.ng@bright.example"),
        (" POPCORN 10pk 100g ", "CC-018", "raj.patel@coastal.example"),
        ("sourdough loaf 800g  ", "HL-207", "sam.lee@harvest.example"),
        (" cheddar cheese 500g", "CC-042", "raj.patel@coastal.example"),
        ("salted  butter 250g ", "GF-115", "kim.tan@greenfield.example"),
        (" GREEK yoghurt 1kg ", "RS-260", "leo.cruz@riverside.example"),
        ("free-range eggs 12pk  ", "GF-119", "kim.tan@greenfield.example"),
        (" basmati rice 5kg ", "RS-271", "leo.cruz@riverside.example"),
    ]
    headers = ["Raw product", "Supplier code", "Contact", "Clean name", "Code prefix", "Label"]
    task = (
        "In D2 type  =PROPER(TRIM(A2))  to clean the name, in E2 type  =LEFT(B2,2)  for the "
        "two-letter prefix, and in F2 type  =D2 & \" (\" & E2 & \")\"  for the label. Fill all three "
        "down. The originals in A:C never change — the tidy columns are built from them."
    )
    for solution in (False, True):
        ws = (
            _add_solution_sheet(wb, "Text formulas soln")
            if solution
            else wb.create_sheet("Text formulas")
        )
        for c, head in enumerate(headers, start=1):
            _header_cell(ws.cell(1, c, head))
        for r, (raw, code, contact) in enumerate(raw_rows, start=2):
            ws.cell(r, 1, raw)
            ws.cell(r, 2, code)
            ws.cell(r, 3, contact)
            if solution:
                ws.cell(r, 4, f"=PROPER(TRIM(A{r}))")
                ws.cell(r, 5, f"=LEFT(B{r},2)")
                ws.cell(r, 6, f'=D{r} & " (" & E{r} & ")"')
        widths = {"A": 26, "B": 14, "C": 28, "D": 24, "E": 12, "F": 28}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
        if not solution:
            _task_note(ws, "H1", "H2", "H2:L6", task)
            ws.column_dimensions["H"].width = 16
            for col in ("I", "J", "K", "L"):
                ws.column_dimensions[col].width = 12


def build_aggregation(wb: openpyxl.Workbook) -> None:
    """Aggregation & precision: a product list beside a count/rounded-average block."""
    agg_rows = [
        ("Snacks", "Stacked Chips 12pk 300g", 120, 3480.0),
        ("Snacks", "Corn Puffs 12pk 150g", 90, 2232.0),
        ("Snacks", "Rice Crackers 24pk 100g", 70, 2625.0),
        ("Snacks", "Popcorn 10pk 100g", 130, 2210.0),
        ("Bakery", "Sourdough Loaf 800g", 85, 3060.0),
        ("Bakery", "Ciabatta Rolls 6pk", 60, 1440.0),
        ("Bakery", "Bagels 6pk", 75, 1875.0),
        ("Dairy", "Cheddar Cheese 500g", 50, 2450.0),
        ("Dairy", "Salted Butter 250g", 55, 1749.0),
        ("Dairy", "Greek Yoghurt 1kg", 65, 1248.0),
        ("Dairy", "Milk 2L", 140, 2100.0),
        ("Produce", "Bananas 1kg", 200, 1000.0),
        ("Produce", "Baby Spinach 200g", 90, 810.0),
        ("Produce", "Cherry Tomatoes 250g", 110, 1320.0),
    ]
    categories = ["Snacks", "Bakery", "Dairy", "Produce"]
    data_headers = ["Category", "Product", "Cases", "Revenue (AUD)"]
    task = (
        "The categories are listed from F2 down. In G2 type  =COUNTIFS($A$2:$A$15,F2)  and fill down "
        "for the line count. In H2 type  =ROUND(AVERAGEIFS($D$2:$D$15,$A$2:$A$15,F2),0)  and fill "
        "down for the whole-dollar average. The $ locks the data ranges as you fill."
    )
    for solution in (False, True):
        ws = (
            _add_solution_sheet(wb, "Aggregation soln")
            if solution
            else wb.create_sheet("Aggregation")
        )
        for c, head in enumerate(data_headers, start=1):
            _header_cell(ws.cell(1, c, head))
        for r, (cat, product, cases, revenue) in enumerate(agg_rows, start=2):
            ws.cell(r, 1, cat)
            ws.cell(r, 2, product)
            ws.cell(r, 3, cases)
            ws.cell(r, 4, revenue).number_format = "#,##0"
        _header_cell(ws.cell(1, 6, "Category"))
        _header_cell(ws.cell(1, 7, "Line count"))
        _header_cell(ws.cell(1, 8, "Avg revenue"))
        for r, cat in enumerate(categories, start=2):
            ws.cell(r, 6, cat)
            if solution:
                ws.cell(r, 7, f"=COUNTIFS($A$2:$A$15,F{r})")
                ws.cell(r, 8, f"=ROUND(AVERAGEIFS($D$2:$D$15,$A$2:$A$15,F{r}),0)").number_format = "#,##0"
        widths = {"A": 12, "B": 26, "C": 8, "D": 15, "F": 12, "G": 12, "H": 14}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
        if not solution:
            _task_note(ws, "F7", "F8", "F8:J12", task)
            for col in ("I", "J"):
                ws.column_dimensions[col].width = 12


def build_dates(wb: openpyxl.Workbook) -> None:
    """Date & time functions: orders with dates and lead times to compute against."""
    date_rows = [
        ("Stacked Chips 12pk 300g", date(2026, 1, 6), 14),
        ("Corn Puffs 12pk 150g", date(2026, 1, 13), 10),
        ("Rice Crackers 24pk 100g", date(2026, 1, 20), 21),
        ("Popcorn 10pk 100g", date(2026, 2, 3), 7),
        ("Sourdough Loaf 800g", date(2026, 2, 10), 3),
        ("Cheddar Cheese 500g", date(2026, 2, 24), 14),
        ("Salted Butter 250g", date(2026, 3, 3), 10),
        ("Greek Yoghurt 1kg", date(2026, 3, 17), 5),
        ("Free-range Eggs 12pk", date(2026, 3, 24), 2),
        ("Basmati Rice 5kg", date(2026, 4, 7), 28),
    ]
    headers = ["Product", "Order date", "Lead days", "Due date", "Month", "Month end", "Age (days)"]
    task = (
        "In D2 type  =B2+C2  for the due date, in E2 type  =TEXT(B2,\"mmm\")  for the month label, "
        "in F2 type  =EOMONTH(B2,0)  for the month-end, and in G2 type  =TODAY()-B2  for the age in "
        "days. Fill down. If a due date shows as a number, format the column as a date."
    )
    date_fmt = "dd mmm yyyy"
    for solution in (False, True):
        ws = (
            _add_solution_sheet(wb, "Dates soln")
            if solution
            else wb.create_sheet("Dates")
        )
        for c, head in enumerate(headers, start=1):
            _header_cell(ws.cell(1, c, head))
        for r, (product, order_date, lead) in enumerate(date_rows, start=2):
            ws.cell(r, 1, product)
            ws.cell(r, 2, order_date).number_format = date_fmt
            ws.cell(r, 3, lead)
            if solution:
                ws.cell(r, 4, f"=B{r}+C{r}").number_format = date_fmt
                ws.cell(r, 5, f'=TEXT(B{r},"mmm")')
                ws.cell(r, 6, f"=EOMONTH(B{r},0)").number_format = date_fmt
                ws.cell(r, 7, f"=TODAY()-B{r}").number_format = "#,##0"
        widths = {"A": 26, "B": 14, "C": 11, "D": 14, "E": 8, "F": 14, "G": 11}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
        if not solution:
            _task_note(ws, "I1", "I2", "I2:M6", task)
            for col in ("I", "J", "K", "L", "M"):
                ws.column_dimensions[col].width = 12


def build_dynamic_arrays(wb: openpyxl.Workbook) -> None:
    """Dynamic arrays: a product list beside spilled UNIQUE/SORT and FILTER results."""
    array_rows = [
        ("Snacks", "Stacked Chips 12pk 300g", 3480.0),
        ("Snacks", "Corn Puffs 12pk 150g", 2232.0),
        ("Snacks", "Rice Crackers 24pk 100g", 2625.0),
        ("Snacks", "Popcorn 10pk 100g", 2210.0),
        ("Bakery", "Sourdough Loaf 800g", 3060.0),
        ("Bakery", "Ciabatta Rolls 6pk", 1440.0),
        ("Bakery", "Bagels 6pk", 1875.0),
        ("Dairy", "Cheddar Cheese 500g", 2450.0),
        ("Dairy", "Salted Butter 250g", 1749.0),
        ("Dairy", "Greek Yoghurt 1kg", 1248.0),
        ("Dairy", "Milk 2L", 2100.0),
        ("Produce", "Bananas 1kg", 1000.0),
        ("Produce", "Baby Spinach 200g", 810.0),
        ("Produce", "Cherry Tomatoes 250g", 1320.0),
    ]
    data_headers = ["Category", "Product", "Revenue (AUD)"]
    task = (
        "In E2 type  =SORT(UNIQUE(A2:A15))  and press Enter — a sorted, de-duplicated category list "
        "spills down. In G2 type  =FILTER(B2:C15,A2:A15=\"Snacks\")  — the Snacks product and revenue "
        "spill into G and H. Don't fill these down; the single formula does it all."
    )
    for solution in (False, True):
        ws = (
            _add_solution_sheet(wb, "Dynamic arrays soln")
            if solution
            else wb.create_sheet("Dynamic arrays")
        )
        for c, head in enumerate(data_headers, start=1):
            _header_cell(ws.cell(1, c, head))
        for r, (cat, product, revenue) in enumerate(array_rows, start=2):
            ws.cell(r, 1, cat)
            ws.cell(r, 2, product)
            ws.cell(r, 3, revenue).number_format = "#,##0"
        _header_cell(ws.cell(1, 5, "Categories"))
        _header_cell(ws.cell(1, 7, "Snacks product"))
        _header_cell(ws.cell(1, 8, "Revenue"))
        if solution:
            # SORT/UNIQUE/FILTER are inherently dynamic; openpyxl can only emit
            # legacy CSE array formulas that Excel strips on load. Left empty here
            # and written natively via Excel COM in _apply_dynamic_array_formulas.
            pass
        widths = {"A": 12, "B": 26, "C": 15, "E": 14, "G": 26, "H": 12}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
        if not solution:
            _task_note(ws, "E4", "E5", "E5:I9", task)


def build_named_ranges(wb: openpyxl.Workbook) -> None:
    """Named ranges: a revenue list to name, then total and average by name."""
    named_rows = [
        ("Stacked Chips 12pk 300g", 3480.0),
        ("Corn Puffs 12pk 150g", 2232.0),
        ("Rice Crackers 24pk 100g", 2625.0),
        ("Popcorn 10pk 100g", 2210.0),
        ("Sourdough Loaf 800g", 3060.0),
        ("Cheddar Cheese 500g", 2450.0),
        ("Salted Butter 250g", 1749.0),
        ("Greek Yoghurt 1kg", 1248.0),
        ("Free-range Eggs 12pk", 2318.0),
        ("Basmati Rice 5kg", 1056.0),
    ]
    data_headers = ["Product", "Revenue (AUD)"]
    task = (
        "Select the revenue figures in B2:B11, click the Name Box (left of the formula bar), type "
        "Revenue and press Enter. Then in D2 type  =SUM(Revenue)  and in D3 type "
        "=AVERAGE(Revenue) . Open Formulas > Name Manager to see the name listed."
    )
    soln_name = "Named ranges soln"
    for solution in (False, True):
        ws = (
            _add_solution_sheet(wb, soln_name)
            if solution
            else wb.create_sheet("Named ranges")
        )
        for c, head in enumerate(data_headers, start=1):
            _header_cell(ws.cell(1, c, head))
        for r, (product, revenue) in enumerate(named_rows, start=2):
            ws.cell(r, 1, product)
            ws.cell(r, 2, revenue).number_format = "#,##0"
        if solution:
            ws.cell(2, 3, "Total")
            ws.cell(2, 4, "=SUM(Revenue)").number_format = "#,##0"
            ws.cell(3, 3, "Average")
            ws.cell(3, 4, "=AVERAGE(Revenue)").number_format = "#,##0"
            ref = f"'{soln_name}'!$B$2:$B$11"
            dn = DefinedName("Revenue", attr_text=ref)
            try:
                wb.defined_names["Revenue"] = dn
            except (TypeError, AttributeError):
                wb.defined_names.append(dn)
        widths = {"A": 26, "B": 15, "C": 10, "D": 12}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
        if not solution:
            _task_note(ws, "D1", "D2", "D2:H6", task)
            for col in ("E", "F", "G", "H"):
                ws.column_dimensions[col].width = 12


def build_power_query(wb: openpyxl.Workbook) -> None:
    """Get & Transform: a deliberately messy export to clean, plus the tidy target.

    openpyxl cannot author a live query, so the starter supplies raw data with the
    classic problems (title row, stray spaces, mixed casing, a blank row, revenue as
    text) and the solution shows the loaded, cleaned result to compare against.
    """
    task = (
        "Click a cell in the block, then Data > From Table/Range. In the editor: Remove Top Rows (the "
        "title), Use First Row as Headers, Trim the text columns, set Revenue to a number, Remove "
        "Blank Rows, then Close & Load. Compare with the 'Power Query soln' tab."
    )
    messy = [
        ("WEEKLY EXPORT \u2014 do not edit above", None, None, None),
        ("Product ", " Category", "Cases", "Revenue "),
        (" stacked chips 12pk 300g", "snacks ", 120, "3480"),
        ("CORN PUFFS 12pk 150g ", " Snacks", 90, "2232"),
        (None, None, None, None),
        (" sourdough loaf 800g", "bakery", 60, "2160"),
        ("cheddar cheese 500g ", " DAIRY", 45, "2205"),
        (" free-range eggs 12pk", "produce ", 80, "1952"),
    ]
    clean = [
        ("Product", "Category", "Cases", "Revenue (AUD)"),
        ("Stacked Chips 12pk 300g", "Snacks", 120, 3480.0),
        ("Corn Puffs 12pk 150g", "Snacks", 90, 2232.0),
        ("Sourdough Loaf 800g", "Bakery", 60, 2160.0),
        ("Cheddar Cheese 500g", "Dairy", 45, 2205.0),
        ("Free-range Eggs 12pk", "Produce", 80, 1952.0),
    ]

    raw = wb.create_sheet("Power Query")
    for r, row in enumerate(messy, start=1):
        for c, val in enumerate(row, start=1):
            if val is not None:
                raw.cell(r, c, val)
    widths = {"A": 26, "B": 14, "C": 9, "D": 14}
    for col, w in widths.items():
        raw.column_dimensions[col].width = w
    _task_note(raw, "F1", "F2", "F2:J6", task)
    for col in ("G", "H", "I", "J"):
        raw.column_dimensions[col].width = 12

    soln = _add_solution_sheet(wb, "Power Query soln")
    for r, row in enumerate(clean, start=1):
        for c, val in enumerate(row, start=1):
            cell = soln.cell(r, c, val)
            if r == 1:
                _header_cell(cell)
            elif isinstance(val, float):
                cell.number_format = "#,##0"
    for col, w in widths.items():
        soln.column_dimensions[col].width = w
    soln.freeze_panes = "A2"


EXTRA_BUILDERS = [
    build_l03_1_slicers,
    build_l04_sumifs,
    build_l05_xlookup,
    build_l06_conditional,
    build_l07_chart,
    build_l08_flashfill,
    build_l09_print,
    build_sort_filter,
    build_data_validation,
    build_logical_iferror,
    build_number_formatting,
    build_text_formulas,
    build_aggregation,
    build_dates,
    build_dynamic_arrays,
    build_named_ranges,
    build_power_query,
]


def write_index(ws: Worksheet, catalogue: list[tuple[str, str, str]]) -> None:
    ws["A1"] = "Excel to a High Standard — Course Workbook"
    ws["A1"].font = Font(bold=True, size=16, color="1A7A4C")
    ws["A3"] = "One file for the whole course. Each lesson has its own tab with the data ready to go."
    ws["A4"] = "How to use each lesson:"
    ws["A4"].font = Font(bold=True)
    for i, step in enumerate([
        "1. Open this workbook and click the lesson's tab (starter data is already set up).",
        "2. Do the task yourself first — the 'Your task' note on each tab tells you what to build.",
        "3. Then copy the Microsoft 365 Copilot prompt from the lesson and watch it do the same task on this tab.",
        "4. Then copy the GitHub Copilot prompt and watch an agent build it from scratch.",
    ], start=5):
        ws.cell(i, 1, step)

    ws["A11"] = "Lesson tabs"
    ws["A11"].font = Font(bold=True)
    ws["A12"] = "Tab"
    ws["B12"] = "Lesson"
    _header_cell(ws["A12"])
    _header_cell(ws["B12"])
    for i, (number, title, sheet) in enumerate(catalogue, start=13):
        ws.cell(i, 1, sheet)
        ws.cell(i, 2, f"Lesson {number} — {title}")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 56


def _reorder_sheets(wb: openpyxl.Workbook, sheet_order: list[str]) -> None:
    """Reorder tabs to follow the course sequence.

    Each starter sheet keeps the tabs the builders created after it (e.g. its
    hidden ``... soln``) as a contiguous group, so groups move together regardless
    of solution-tab naming. Leading non-lesson sheets (Index) stay first; any tab
    not named in ``sheet_order`` is preserved at the end.
    """
    order_set = set(sheet_order)
    head: list = []
    groups: dict[str, list] = {}
    current: str | None = None
    for ws in wb._sheets:
        if ws.title in order_set:
            current = ws.title
            groups[current] = [ws]
        elif current is None:
            head.append(ws)
        else:
            groups[current].append(ws)

    reordered = list(head)
    for name in sheet_order:
        reordered.extend(groups.pop(name, []))
    for leftover in groups.values():
        reordered.extend(leftover)
    wb._sheets = reordered



DYNAMIC_ARRAY_FORMULAS: list[tuple[str, str, str]] = [
    ("Dynamic arrays soln", "E2", "=SORT(UNIQUE(A2:A15))"),
    ("Dynamic arrays soln", "G2", '=FILTER(B2:C15,A2:A15="Snacks")'),
]


def _apply_dynamic_array_formulas(path: Path) -> bool:
    """Write spilling dynamic-array formulas natively via Excel COM.

    openpyxl can only emit legacy CSE array formulas, which Excel strips on load
    for inherently-dynamic functions (SORT/UNIQUE/FILTER). Excel's own Formula2
    property enters them as valid spilling arrays that open without repair.
    Returns True when the COM pass ran, False when Excel/pywin32 is unavailable.
    """
    try:
        import win32com.client as win32
    except ImportError:
        return False

    xl_openxml_workbook = 51
    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(str(path))
        for sheet_name, cell, formula in DYNAMIC_ARRAY_FORMULAS:
            wb.Worksheets(sheet_name).Range(cell).Formula2 = formula
        wb.SaveAs(str(path), FileFormat=xl_openxml_workbook)
        wb.Close(SaveChanges=False)
    finally:
        excel.Quit()
    return True


def build() -> None:
    catalogue = _course_catalogue()
    wb = openpyxl.Workbook()
    index = wb.active
    index.title = "Index"
    write_index(index, catalogue)

    for lesson in LESSONS:
        ws = wb.create_sheet(lesson.sheet)
        write_lesson(ws, lesson)
        if lesson.rate_named:
            ref = f"'{lesson.sheet}'!$F$1"
            dn = DefinedName("Rate", attr_text=ref)
            try:
                wb.defined_names["Rate"] = dn
            except (TypeError, AttributeError):
                wb.defined_names.append(dn)

    for builder in EXTRA_BUILDERS:
        builder(wb)

    _reorder_sheets(wb, [sheet for _, _, sheet in catalogue])

    wb.save(OUT)
    if not _apply_dynamic_array_formulas(OUT):
        print("WARNING: Excel/pywin32 unavailable — dynamic-array soln formulas not written.")
    print(f"Wrote {len(wb.sheetnames)} sheets ({', '.join(wb.sheetnames)}) to {OUT.name}")


if __name__ == "__main__":
    build()
